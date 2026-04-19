"""Geração de relatórios Excel — SPARTA AGENTE IA."""
import logging
from datetime import datetime
from pathlib import Path

log = logging.getLogger(__name__)


def gerar_excel(
    registros: list,
    destino: Path | None = None,
    titulo: str = "Relatório SPARTA AGENTE IA",
) -> Path:
    """
    Gera planilha Excel com os registros de análises fornecidos.
    Retorna o caminho do arquivo gerado.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment, Font, PatternFill, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl não instalado. Execute: pip install openpyxl")

    if destino is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        destino = Path(".") / f"relatorio_sparta_{ts}.xlsx"

    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alertas"

    # Paleta
    AMARELO  = "FFD000"
    PRETO    = "0F0F0F"
    CINZA    = "D0D0D0"
    VERMELHO = "FF4444"
    LARANJA  = "FF8800"
    VERDE    = "22AA55"
    BRANCO   = "FFFFFF"

    COR_NIVEL = {
        "critico":   VERMELHO,
        "suspeito":  LARANJA,
        "atencao":   "FFCC00",
        "sem_risco": VERDE,
    }

    thin = Side(style="thin", color="AAAAAA")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:K1")
    cel = ws["A1"]
    cel.value = titulo
    cel.font  = Font(name="Segoe UI", bold=True, size=14, color=BRANCO)
    cel.fill  = PatternFill("solid", fgColor=PRETO)
    cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Gerado em
    ws.merge_cells("A2:K2")
    cel2 = ws["A2"]
    cel2.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}    Total de registros: {len(registros)}"
    cel2.font  = Font(name="Segoe UI", size=9, color="555555")
    cel2.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Cabeçalhos
    cabecalhos = [
        "ID", "Data/Hora", "Câmera", "Fase",
        "Nível de Risco", "Alerta", "Confiança",
        "Comportamentos Detectados", "Ação Recomendada",
        "Tokens Entrada", "Tokens Saída",
    ]
    linha_cab = 3
    for col, texto in enumerate(cabecalhos, 1):
        c = ws.cell(row=linha_cab, column=col, value=texto)
        c.font      = Font(name="Segoe UI", bold=True, size=10, color=PRETO)
        c.fill      = PatternFill("solid", fgColor=AMARELO)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borda
    ws.row_dimensions[linha_cab].height = 22

    # Dados
    for i, reg in enumerate(registros, 1):
        linha = linha_cab + i

        comportamentos = reg.get("comportamentos", "")
        if isinstance(comportamentos, list):
            comportamentos = "; ".join(comportamentos)

        ts_raw = reg.get("timestamp_analise") or reg.get("created_at") or ""
        try:
            ts_fmt = datetime.fromisoformat(ts_raw.replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M:%S")
        except Exception:
            ts_fmt = ts_raw

        nivel   = (reg.get("nivel_risco") or "sem_risco").lower()
        alerta  = "SIM" if reg.get("alerta") else "Não"
        conf    = reg.get("confianca", 0)
        conf_p  = f"{float(conf):.0%}" if conf is not None else "-"

        valores = [
            reg.get("id", i),
            ts_fmt,
            reg.get("camera_id", ""),
            reg.get("fase_processo", ""),
            nivel.upper(),
            alerta,
            conf_p,
            comportamentos[:500] if comportamentos else "",
            reg.get("acao_recomendada", "")[:300],
            reg.get("tokens_entrada", 0) or 0,
            reg.get("tokens_saida", 0) or 0,
        ]

        cor_fundo = COR_NIVEL.get(nivel, CINZA)
        fill_nivel = PatternFill("solid", fgColor=cor_fundo)

        for col, valor in enumerate(valores, 1):
            c = ws.cell(row=linha, column=col, value=valor)
            c.font      = Font(name="Segoe UI", size=9)
            c.border    = borda
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 5:  # coluna Nível de Risco
                c.fill = fill_nivel
                c.font = Font(name="Segoe UI", bold=True, size=9,
                              color=BRANCO if nivel in ("critico", "suspeito") else PRETO)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif i % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F5F5F5")

        ws.row_dimensions[linha].height = 18

    # Larguras das colunas
    larguras = [6, 18, 14, 12, 14, 8, 10, 60, 50, 13, 12]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    # Aba de resumo
    ws2 = wb.create_sheet("Resumo")
    _preencher_resumo(ws2, registros, AMARELO, PRETO, borda, COR_NIVEL)

    wb.save(str(destino))
    log.info("Relatório Excel gerado: %s (%d registros)", destino.name, len(registros))
    return destino


def _preencher_resumo(ws, registros, AMARELO, PRETO, borda, COR_NIVEL):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    contagem = {"critico": 0, "suspeito": 0, "atencao": 0, "sem_risco": 0}
    tokens_total_in = 0
    tokens_total_out = 0

    for reg in registros:
        nivel = (reg.get("nivel_risco") or "sem_risco").lower()
        contagem[nivel] = contagem.get(nivel, 0) + 1
        tokens_total_in  += int(reg.get("tokens_entrada", 0) or 0)
        tokens_total_out += int(reg.get("tokens_saida", 0) or 0)

    ws["A1"] = "Resumo por Nível de Risco"
    ws["A1"].font = Font(name="Segoe UI", bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=PRETO)

    ws["A2"] = "Nível"
    ws["B2"] = "Quantidade"
    for c in (ws["A2"], ws["B2"]):
        c.font  = Font(name="Segoe UI", bold=True, size=10)
        c.fill  = PatternFill("solid", fgColor=AMARELO)
        c.border = borda
        c.alignment = Alignment(horizontal="center")

    for i, (nivel, qtd) in enumerate(contagem.items(), 3):
        ws.cell(row=i, column=1, value=nivel.upper()).border = borda
        ws.cell(row=i, column=2, value=qtd).border = borda

    linha_tok = len(contagem) + 4
    ws.cell(row=linha_tok,     column=1, value="Tokens Entrada Total").font = Font(bold=True, name="Segoe UI", size=9)
    ws.cell(row=linha_tok,     column=2, value=tokens_total_in)
    ws.cell(row=linha_tok + 1, column=1, value="Tokens Saída Total").font   = Font(bold=True, name="Segoe UI", size=9)
    ws.cell(row=linha_tok + 1, column=2, value=tokens_total_out)
    ws.cell(row=linha_tok + 2, column=1, value="Total de Análises").font    = Font(bold=True, name="Segoe UI", size=9)
    ws.cell(row=linha_tok + 2, column=2, value=len(registros))

    try:
        chart = BarChart()
        chart.type  = "col"
        chart.title = "Distribuição de Alertas"
        chart.y_axis.title = "Quantidade"
        data = Reference(ws, min_col=2, min_row=2, max_row=2 + len(contagem))
        cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(contagem))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, "D2")
    except Exception:
        pass


def gerar_excel_do_db(
    data_inicio: str | None = None,
    data_fim: str | None = None,
    camera_id: str | None = None,
    destino: Path | None = None,
) -> Path:
    """Busca registros no DB e gera Excel. Filtros opcionais por data e câmera."""
    try:
        import db as _db
        registros = _db.buscar_analises_filtradas(
            data_inicio=data_inicio,
            data_fim=data_fim,
            camera_id=camera_id,
        )
    except Exception as exc:
        log.error("Erro ao buscar registros para relatório: %s", exc)
        registros = []

    titulo = "Relatório SPARTA AGENTE IA"
    if data_inicio or data_fim:
        titulo += f" — {data_inicio or ''} a {data_fim or ''}"
    return gerar_excel(registros, destino=destino, titulo=titulo)
